"""
financial_report.py
────────────────────────────────────────────────────────────────
公司财务报表计算引擎

支持三大报表：
  1. build_pnl()           — 损益表 (P&L Statement)
  2. build_ar_summary()    — 应收账款摘要 (AR Summary)
  3. build_expense_analysis() — 支出分析 (Expense Analysis)

导出：
  4. build_pnl_pdf()       — ReportLab PDF
  5. build_pnl_excel()     — openpyxl Excel
"""

import io
from datetime import datetime, date
from decimal import Decimal

from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_

from models.invoice import Invoice, Payment, InvoiceStatus
from models.expense import Expense, ExpenseCategory, ExpenseStatus
from models.commission import CommissionStatement, IR56MStatement, CommissionStatus
from models.company import Company

# ──────────────────────────────────────────────────────────────
# 工具函数
# ──────────────────────────────────────────────────────────────

def _fiscal_year_dates(fiscal_year: str) -> tuple[date, date]:
    """
    将 '2024-25' 转换为起止日期：
    2024-04-01 ~ 2025-03-31
    """
    try:
        start_yr = int(fiscal_year.split("-")[0])
    except Exception:
        start_yr = datetime.utcnow().year
    start = date(start_yr, 4, 1)
    end   = date(start_yr + 1, 3, 31)
    return start, end


def _to_float(val) -> float:
    if val is None:
        return 0.0
    return float(val)


def _fmt(val: float, decimals: int = 2) -> str:
    return f"{val:,.{decimals}f}"


def get_current_fiscal_year() -> str:
    """返回当前香港财政年度字符串，如 '2024-25'"""
    today = date.today()
    if today.month >= 4:
        return f"{today.year}-{str(today.year + 1)[-2:]}"
    else:
        return f"{today.year - 1}-{str(today.year)[-2:]}"


def _estimate_profits_tax(profit: float) -> dict:
    """法团利得税：首200万 8.25%，超出 16.5%"""
    if profit <= 0:
        return {"rate_desc": "无须征税", "estimated_tax": 0.0, "effective_rate": 0.0}
    tier1_limit = 2_000_000.0
    if profit <= tier1_limit:
        tax = profit * 0.0825
        desc = "8.25%（法团两级制首级）"
    else:
        tax = tier1_limit * 0.0825 + (profit - tier1_limit) * 0.165
        desc = "8.25% / 16.5%（法团两级制）"
    return {"rate_desc": desc, "estimated_tax": round(tax, 2), "effective_rate": round(tax / profit * 100, 2)}

def _estimate_unincorporated_profits_tax(profit: float) -> dict:
    """非法团利得税：首200万 7.5%，超出 15%"""
    if profit <= 0:
        return {"rate_desc": "无须征税", "estimated_tax": 0.0, "effective_rate": 0.0, "formula": "利润为负，免税"}
    tier1_limit = 2_000_000.0
    if profit <= tier1_limit:
        tax = profit * 0.075
        desc = "7.5%（非法团两级制首级）"
        formula = f"{profit:,.0f} × 7.5%"
    else:
        tax = tier1_limit * 0.075 + (profit - tier1_limit) * 0.15
        desc = "7.5% / 15%（非法团两级制）"
        formula = f"2,000,000 × 7.5% + {(profit - tier1_limit):,.0f} × 15%"
    return {"rate_desc": desc, "estimated_tax": round(tax, 2), "effective_rate": round(tax / profit * 100, 2), "formula": formula}

def calculate_salaries_tax(net_chargeable_income: float, total_net_income: float) -> dict:
    if total_net_income <= 0:
        return {"tax": 0.0, "method": "无须征税"}
    standard_tax = total_net_income * 0.15
    if net_chargeable_income <= 0:
        return {"tax": 0.0, "method": "免税额已全数抵扣"}
    
    progressive_tax = 0.0
    rem = net_chargeable_income
    for limit, rate in [(50000, 0.02), (50000, 0.06), (50000, 0.10), (50000, 0.14)]:
        if rem > limit:
            progressive_tax += limit * rate
            rem -= limit
        else:
            progressive_tax += rem * rate
            rem = 0
            break
    if rem > 0:
        progressive_tax += rem * 0.17
        
    if progressive_tax < standard_tax:
        return {"tax": progressive_tax, "method": "薪俸累进税率 (2%-17%)"}
    else:
        return {"tax": standard_tax, "method": "15% 标准税率退保"}

def evaluate_personal_assessment(profit: float, profile) -> dict:
    res = {}
    mpf = float(profile.mpf_self_contribution or 0)
    other = float(profile.other_deductions or 0)
    my_net_income = max(profit - mpf - other, 0.0)
    
    child = int(profile.children_count or 0)
    p60 = int(profile.dependent_parents_60 or 0)
    p55 = int(profile.dependent_parents_55 or 0)
    
    basic_allw, married_allw, child_allw = 132000, 264000, child * 130000
    parent_allw = p60 * 50000 + p55 * 25000
    marital_status = profile.marital_status or 'single'
    
    if marital_status == 'single':
        total_allw = basic_allw + child_allw + parent_allw
        nci = max(my_net_income - total_allw, 0.0)
        t = calculate_salaries_tax(nci, my_net_income)
        res['strategy'] = 'single_personal_assessment'
        res['tax'] = t['tax']
        formula_str = f"应课税入息 = 利润 - 免税额({total_allw:,.0f}) = {nci:,.0f} | 适用累进率"
        res['breakdown'] = {'method': '单身个人入息课税 (' + t['method'] + ')', 'nci': nci, 'allowance': total_allw, 'formula': formula_str}
        return res
        
    spouse_inc = float(profile.spouse_net_income or 0)
    
    # 方案1: 分开评税
    tot_allw_sep = basic_allw + child_allw + parent_allw
    my_nci_sep = max(my_net_income - tot_allw_sep, 0.0)
    my_tax_sep = calculate_salaries_tax(my_nci_sep, my_net_income)['tax']
    spouse_nci_sep = max(spouse_inc - basic_allw, 0.0)
    spouse_tax_sep = calculate_salaries_tax(spouse_nci_sep, spouse_inc)['tax']
    sep_total = my_tax_sep + spouse_tax_sep
    
    # 方案2: 合并评税
    joint_net = my_net_income + spouse_inc
    joint_allw = married_allw + child_allw + parent_allw
    joint_nci = max(joint_net - joint_allw, 0.0)
    joint_t = calculate_salaries_tax(joint_nci, joint_net)
    joint_total = joint_t['tax']
    
    if sep_total <= joint_total:
        res['strategy'] = 'separate_assessment'
        res['tax'] = my_tax_sep
        res['family_total_tax'] = sep_total
        formula_str = f"分开评税：扣除各自强积金后，扣减各自独立免税额，单独适用累进率"
        res['breakdown'] = {'method': '夫妇分开评税更有利', 'my_tax': my_tax_sep, 'spouse_tax': spouse_tax_sep, 'formula': formula_str}
    else:
        res['strategy'] = 'joint_assessment'
        # 计算在合并中我应当分担的比例或者直接报整个家庭税
        res['tax'] = max(joint_total - spouse_tax_sep, 0.0)
        res['family_total_tax'] = joint_total
        formula_str = f"全家应课税入息 = 家族总入息 - 总免税额({joint_allw:,.0f}) = {joint_nci:,.0f} | 适用累进率"
        res['breakdown'] = {'method': '合并评税更有利 (' + joint_t['method'] + ')', 'joint_tax': joint_total, 'formula': formula_str}
    return res



# ──────────────────────────────────────────────────────────────
# 1. 损益表
# ──────────────────────────────────────────────────────────────

def build_pnl(db: Session, company_id: str, fiscal_year: str) -> dict:
    """
    生成损益表数据字典。

    收入来源：
      - invoices（已付款 paid / partial）
      - commission_statements（已确认）
      - ir56m_statements（已确认，若存在则优先替代月结单）

    支出来源：
      - expenses（已确认 confirmed），按分类分组
    """
    start, end = _fiscal_year_dates(fiscal_year)

    company = db.query(Company).filter(Company.id == company_id).first()
    business_mode = company.business_mode if company else "trading_sme"
    base_currency = company.base_currency if company else "HKD"

    # ── 1. 发票收入 ────────────────────────────────────────────
    invoice_income = 0.0
    invoice_rows   = []

    invoices = db.query(Invoice).filter(
        Invoice.company_id == company_id,
        Invoice.status.in_([InvoiceStatus.paid, InvoiceStatus.partial]),
        func.date(Invoice.issue_date) >= start,
        func.date(Invoice.issue_date) <= end,
    ).all()

    for inv in invoices:
        paid = _to_float(inv.paid_amount)
        invoice_income += paid
        invoice_rows.append({
            "invoice_number": inv.invoice_number,
            "client_name":    inv.client_name,
            "issue_date":     inv.issue_date.strftime("%Y-%m-%d") if inv.issue_date else "",
            "amount":         paid,
            "currency":       inv.currency,
        })

    # ── 2. 佣金收入 ────────────────────────────────────────────
    commission_income     = 0.0
    commission_source     = "monthly"   # "ir56m" | "monthly"
    commission_rows       = []

    # 优先检查已确认的 IR56M 年结单
    ir56m = db.query(IR56MStatement).filter(
        IR56MStatement.company_id == company_id,
        IR56MStatement.fiscal_year == fiscal_year,
        IR56MStatement.status == "confirmed",
    ).first()

    if ir56m:
        commission_income = _to_float(ir56m.total_income)
        commission_source = "ir56m"
        period_str = f"{ir56m.period_start} 到 {ir56m.period_end}" if (ir56m.period_start and ir56m.period_end) else fiscal_year
        commission_rows.append({
            "source":       "IR56M 年度申报表",
            "period":       period_str,
            "amount":       commission_income,
            "note":         "来自 IR56M 官方年度申报（优先级高于月结单）",
        })
    else:
        stmts = db.query(CommissionStatement).filter(
            CommissionStatement.company_id == company_id,
            CommissionStatement.fiscal_year == fiscal_year,
            CommissionStatement.status == CommissionStatus.confirmed,
        ).all()
        for s in stmts:
            amt = _to_float(s.total_taxable_income)
            commission_income += amt
            commission_rows.append({
                "source":  f"月结单 {s.statement_month or ''}",
                "period":  s.statement_month or "",
                "amount":  amt,
            })

    # ── 3. 总收入 ──────────────────────────────────────────────
    total_income = invoice_income + commission_income

    # ── 4. 支出（已确认，按分类分组）──────────────────────────
    expenses_q = db.query(
        ExpenseCategory.code,
        ExpenseCategory.name_zh,
        ExpenseCategory.name_en,
        ExpenseCategory.hk_tax_deductible,
        func.sum(Expense.amount_hkd).label("total_hkd"),
        func.count(Expense.id).label("count"),
    ).join(
        Expense, Expense.category_id == ExpenseCategory.id, isouter=True
    ).filter(
        Expense.company_id == company_id,
        Expense.status == ExpenseStatus.confirmed,
        Expense.fiscal_year == fiscal_year,
    ).group_by(
        ExpenseCategory.id
    ).all()

    expense_categories = []
    total_expense        = 0.0
    total_tax_deductible = 0.0

    for row in expenses_q:
        amt = _to_float(row.total_hkd)
        if amt == 0:
            continue
        is_deductible = row.hk_tax_deductible in ("yes", "partial", "depreciation")
        total_expense += amt
        if is_deductible:
            total_tax_deductible += amt
        expense_categories.append({
            "code":            row.code,
            "name_zh":         row.name_zh,
            "name_en":         row.name_en or row.code,
            "hk_tax_deductible": row.hk_tax_deductible,
            "total_hkd":       round(amt, 2),
            "count":           row.count,
            "deductible_pct":  round(amt / total_expense * 100, 1) if total_expense > 0 else 0,
        })

    # 排序：金额降序
    expense_categories.sort(key=lambda x: -x["total_hkd"])

    # 未分类支出（category_id is NULL）
    uncategorised_q = db.query(func.sum(Expense.amount_hkd), func.count(Expense.id)).filter(
        Expense.company_id == company_id,
        Expense.status == ExpenseStatus.confirmed,
        Expense.fiscal_year == fiscal_year,
        Expense.category_id.is_(None),
    ).one()
    uncategorised_amt = _to_float(uncategorised_q[0])
    if uncategorised_amt > 0:
        total_expense += uncategorised_amt
        expense_categories.append({
            "code":              "未分类",
            "name_zh":           "未分类支出",
            "name_en":           "Uncategorised",
            "hk_tax_deductible": "review",
            "total_hkd":         round(uncategorised_amt, 2),
            "count":             uncategorised_q[1] or 0,
            "deductible_pct":    0,
        })

    # ── 5. 利润与税务 ──────────────────────────────────────────
    gross_profit      = total_income - total_expense
    assessable_profit = total_income - total_tax_deductible
    ap = max(assessable_profit, 0)
    
    is_unlimited = company and company.company_legal_type in ("unlimited", "sole_prop")
    
    if is_unlimited:
        from models.company import CompanyTaxProfile
        profile = db.query(CompanyTaxProfile).filter(CompanyTaxProfile.company_id == company_id).first()
        if not profile:
            profile = CompanyTaxProfile(company_id=company_id)
            
        unincorp_tax = _estimate_unincorporated_profits_tax(ap)
        pa_res = evaluate_personal_assessment(ap, profile)
        
        # Determine the best logic
        cand_pa_tax = pa_res['tax']
        cand_un_tax = unincorp_tax['estimated_tax']
        
        is_pa_better = cand_pa_tax < cand_un_tax
        best_tax = cand_pa_tax if is_pa_better else cand_un_tax
        
        tax_info = {
            "is_unlimited": True,
            "estimated_tax": round(best_tax, 2),
            "rate_desc": pa_res['breakdown']['method'] if is_pa_better else unincorp_tax['rate_desc'],
            "effective_rate": round(best_tax / ap * 100, 2) if ap > 0 else 0,
            
            "options": {
                "unincorporated_profits_tax": unincorp_tax,
                "personal_assessment": pa_res
            },
            "recommendation": "personal_assessment" if is_pa_better else "profits_tax"
        }
    else:
        ti = _estimate_profits_tax(ap)
        tax_info = {
            "is_unlimited": False,
            "estimated_tax": ti['estimated_tax'],
            "rate_desc": ti['rate_desc'],
            "effective_rate": ti['effective_rate']
        }
        
    net_profit = gross_profit - tax_info["estimated_tax"]

    return {
        "fiscal_year":          fiscal_year,
        "period_start":         start.isoformat(),
        "period_end":           end.isoformat(),
        "base_currency":        base_currency,
        "business_mode":        business_mode,
        "generated_at":         datetime.utcnow().isoformat(),

        # 收入区块
        "invoice_income":       round(invoice_income, 2),
        "commission_income":    round(commission_income, 2),
        "commission_source":    commission_source,
        "total_income":         round(total_income, 2),
        "invoice_rows":         invoice_rows,
        "commission_rows":      commission_rows,

        # 支出区块
        "total_expense":        round(total_expense, 2),
        "total_tax_deductible": round(total_tax_deductible, 2),
        "expense_categories":   expense_categories,

        # 利润区块
        "gross_profit":         round(gross_profit, 2),
        "assessable_profit":    round(assessable_profit, 2),
        "tax_info":             tax_info,
        "net_profit":           round(net_profit, 2),
    }


# ──────────────────────────────────────────────────────────────
# 2. 应收账款摘要
# ──────────────────────────────────────────────────────────────

def build_ar_summary(db: Session, company_id: str, fiscal_year: str) -> dict:
    """按客户汇总应收账款状态。"""
    start, end = _fiscal_year_dates(fiscal_year)

    invoices = db.query(Invoice).filter(
        Invoice.company_id == company_id,
        Invoice.status.in_([
            InvoiceStatus.sent, InvoiceStatus.partial,
            InvoiceStatus.overdue, InvoiceStatus.paid,
        ]),
        func.date(Invoice.issue_date) >= start,
        func.date(Invoice.issue_date) <= end,
    ).all()

    today = date.today()
    client_map: dict[str, dict] = {}

    for inv in invoices:
        cname = inv.client_name or "未知客户"
        if cname not in client_map:
            client_map[cname] = {
                "client_name":    cname,
                "invoice_count":  0,
                "total_billed":   0.0,
                "total_paid":     0.0,
                "balance_due":    0.0,
                "overdue_amount": 0.0,
            }
        row = client_map[cname]
        total = _to_float(inv.total_amount)
        paid  = _to_float(inv.paid_amount)
        bal   = total - paid

        row["invoice_count"] += 1
        row["total_billed"]  += total
        row["total_paid"]    += paid
        row["balance_due"]   += bal

        # 逾期判断
        if inv.due_date and inv.due_date.date() < today and bal > 0:
            row["overdue_amount"] += bal

    clients = sorted(client_map.values(), key=lambda x: -x["balance_due"])
    for c in clients:
        c["total_billed"]   = round(c["total_billed"], 2)
        c["total_paid"]     = round(c["total_paid"], 2)
        c["balance_due"]    = round(c["balance_due"], 2)
        c["overdue_amount"] = round(c["overdue_amount"], 2)

    total_billed   = sum(c["total_billed"]   for c in clients)
    total_paid     = sum(c["total_paid"]     for c in clients)
    total_balance  = sum(c["balance_due"]    for c in clients)
    total_overdue  = sum(c["overdue_amount"] for c in clients)

    return {
        "fiscal_year":  fiscal_year,
        "period_start": start.isoformat(),
        "period_end":   end.isoformat(),
        "clients":      clients,
        "summary": {
            "total_billed":  round(total_billed, 2),
            "total_paid":    round(total_paid, 2),
            "total_balance": round(total_balance, 2),
            "total_overdue": round(total_overdue, 2),
            "collection_rate": round(total_paid / total_billed * 100, 1) if total_billed > 0 else 0,
        },
    }


# ──────────────────────────────────────────────────────────────
# 3. 支出分析
# ──────────────────────────────────────────────────────────────

def build_expense_analysis(db: Session, company_id: str, fiscal_year: str) -> dict:
    """
    按分类 + 按月份 的支出双维度分析。
    """
    start, end = _fiscal_year_dates(fiscal_year)

    # 按分类汇总（已确认）
    cat_rows = db.query(
        ExpenseCategory.code,
        ExpenseCategory.name_zh,
        ExpenseCategory.hk_tax_deductible,
        func.sum(Expense.amount_hkd).label("total"),
        func.count(Expense.id).label("cnt"),
    ).join(
        Expense, Expense.category_id == ExpenseCategory.id, isouter=True
    ).filter(
        Expense.company_id == company_id,
        Expense.status == ExpenseStatus.confirmed,
        Expense.fiscal_year == fiscal_year,
    ).group_by(ExpenseCategory.id).all()

    categories = []
    grand_total = 0.0
    for r in cat_rows:
        amt = _to_float(r.total)
        grand_total += amt
        categories.append({
            "code":              r.code,
            "name_zh":          r.name_zh,
            "hk_tax_deductible": r.hk_tax_deductible,
            "total":            round(amt, 2),
            "count":            r.cnt or 0,
        })
    categories.sort(key=lambda x: -x["total"])
    for c in categories:
        c["pct"] = round(c["total"] / grand_total * 100, 1) if grand_total > 0 else 0

    # 按月份汇总（财年内每月）
    monthly_rows = db.query(
        func.strftime("%Y-%m", Expense.receipt_date).label("month"),
        func.sum(Expense.amount_hkd).label("total"),
        func.count(Expense.id).label("cnt"),
    ).filter(
        Expense.company_id == company_id,
        Expense.status == ExpenseStatus.confirmed,
        Expense.fiscal_year == fiscal_year,
        Expense.receipt_date.isnot(None),
    ).group_by("month").order_by("month").all()

    monthly = [
        {"month": r.month, "total": round(_to_float(r.total), 2), "count": r.cnt or 0}
        for r in monthly_rows
    ]

    # 可扣税汇总
    deductible_total = sum(
        c["total"] for c in categories
        if c["hk_tax_deductible"] in ("yes", "partial", "depreciation")
    )
    non_deductible_total = grand_total - deductible_total

    return {
        "fiscal_year":        fiscal_year,
        "period_start":       start.isoformat(),
        "period_end":         end.isoformat(),
        "grand_total":        round(grand_total, 2),
        "deductible_total":   round(deductible_total, 2),
        "non_deductible_total": round(non_deductible_total, 2),
        "deductible_pct":     round(deductible_total / grand_total * 100, 1) if grand_total > 0 else 0,
        "categories":         categories,
        "monthly":            monthly,
    }


# ──────────────────────────────────────────────────────────────
# 4. PDF 导出（损益表）
# ──────────────────────────────────────────────────────────────

def build_pnl_pdf(pnl: dict, company_data: dict) -> bytes:
    """生成损益表 PDF，风格与现有发票 PDF 保持一致。"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable,
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    COLOR_NAVY   = colors.HexColor("#1E3A8A")
    COLOR_LIGHT  = colors.HexColor("#EFF6FF")
    COLOR_BORDER = colors.HexColor("#CBD5E1")
    COLOR_GREEN  = colors.HexColor("#166534")
    COLOR_RED    = colors.HexColor("#991B1B")
    COLOR_MUTED  = colors.HexColor("#64748B")

    # 字体
    font_name = "Helvetica"
    font_candidates = [
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/STHeiti Light.ttc",
        "/Library/Fonts/Arial Unicode MS.ttf",
    ]
    for fp in font_candidates:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("ChineseFont", fp))
                font_name = "ChineseFont"
            except Exception:
                pass
            break

    def S(name, **kwargs):
        return ParagraphStyle(name, fontName=font_name, **kwargs)

    styles = {
        "title":      S("title",   fontSize=20, textColor=COLOR_NAVY,  alignment=TA_LEFT,  leading=26),
        "subtitle":   S("sub",     fontSize=11, textColor=COLOR_MUTED,  alignment=TA_LEFT,  leading=16),
        "section":    S("sec",     fontSize=11, textColor=COLOR_NAVY,   leading=16),
        "body":       S("body",    fontSize=9.5, textColor=colors.black,  leading=14),
        "right":      S("right",   fontSize=9.5, textColor=colors.black,  leading=14, alignment=TA_RIGHT),
        "th":         S("th",      fontSize=9.5, textColor=colors.white,  leading=13, alignment=TA_LEFT),
        "th_r":       S("th_r",    fontSize=9.5, textColor=colors.white,  leading=13, alignment=TA_RIGHT),
        "total_lbl":  S("tlbl",    fontSize=11, textColor=COLOR_NAVY,   leading=16, alignment=TA_LEFT),
        "total_val":  S("tval",    fontSize=11, textColor=COLOR_NAVY,   leading=16, alignment=TA_RIGHT),
        "net_lbl":    S("nlbl",    fontSize=13, textColor=COLOR_NAVY,   leading=18, alignment=TA_LEFT),
        "net_val":    S("nval",    fontSize=13, textColor=COLOR_GREEN,  leading=18, alignment=TA_RIGHT),
        "muted":      S("muted",   fontSize=8.5, textColor=COLOR_MUTED,  leading=12),
    }

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=15*mm, leftMargin=15*mm,
        topMargin=18*mm, bottomMargin=18*mm,
    )
    pw = A4[0] - 30*mm
    story = []

    # ── 封面标题 ─────────────────────────────────────────────────
    comp_name = company_data.get("name_zh", "") or company_data.get("name_en", "")
    story.append(Paragraph(f"<b>{comp_name}</b>", styles["title"]))
    story.append(Paragraph(
        f"損益表 / Profit & Loss Statement  ·  {pnl['fiscal_year']} 財政年度",
        styles["subtitle"]
    ))
    story.append(Paragraph(
        f"期間：{pnl['period_start']} 至 {pnl['period_end']}  ·  "
        f"製表日期：{datetime.utcnow().strftime('%Y-%m-%d')}",
        styles["muted"]
    ))
    story.append(Spacer(1, 3*mm))
    story.append(HRFlowable(width="100%", thickness=1.5, color=COLOR_NAVY))
    story.append(Spacer(1, 5*mm))

    ccy = pnl.get("base_currency", "HKD")

    def make_table(data, col_widths, header_rows=1):
        t = Table(data, colWidths=col_widths, repeatRows=header_rows)
        style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, header_rows - 1), COLOR_NAVY),
            ("TEXTCOLOR",  (0, 0), (-1, header_rows - 1), colors.white),
            ("ROWBACKGROUNDS", (0, header_rows), (-1, -1), [colors.white, COLOR_LIGHT]),
            ("GRID",       (0, 0), (-1, -1), 0.3, COLOR_BORDER),
            ("VALIGN",     (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
        ])
        t.setStyle(style)
        return t

    # ── 收入区块 ─────────────────────────────────────────────────
    story.append(Paragraph("<b>一、收入 / Income</b>", styles["section"]))
    story.append(Spacer(1, 2*mm))

    income_data = [[
        Paragraph("<b>項目 / Item</b>", styles["th"]),
        Paragraph("<b>金額 / Amount</b>", styles["th_r"]),
    ]]
    if pnl["invoice_income"] > 0 or pnl.get("business_mode") != "insurance_agent":
        income_data.append([
            Paragraph("發票收入 / Invoice Revenue", styles["body"]),
            Paragraph(f"{ccy} {_fmt(pnl['invoice_income'])}", styles["right"]),
        ])
    if pnl["commission_income"] > 0:
        src_note = "（IR56M年度報表）" if pnl["commission_source"] == "ir56m" else "（月結單彙總）"
        income_data.append([
            Paragraph(f"佣金收入 / Commission Income {src_note}", styles["body"]),
            Paragraph(f"{ccy} {_fmt(pnl['commission_income'])}", styles["right"]),
        ])
    income_data.append([
        Paragraph("<b>合計收入 / Total Income</b>", styles["total_lbl"]),
        Paragraph(f"<b>{ccy} {_fmt(pnl['total_income'])}</b>", styles["total_val"]),
    ])
    story.append(make_table(income_data, [pw * 0.65, pw * 0.35]))
    story.append(Spacer(1, 5*mm))

    # ── 支出区块 ─────────────────────────────────────────────────
    story.append(Paragraph("<b>二、支出 / Expenses</b>", styles["section"]))
    story.append(Spacer(1, 2*mm))

    exp_data = [[
        Paragraph("<b>支出類別 / Category</b>", styles["th"]),
        Paragraph("<b>扣稅資格</b>", styles["th"]),
        Paragraph("<b>金額 / Amount</b>", styles["th_r"]),
    ]]
    for cat in pnl["expense_categories"]:
        deduct_label = {
            "yes":          "✓ 全額",
            "partial":      "◑ 部份",
            "depreciation": "↓ 折舊",
            "no":           "✗ 不可",
            "review":       "？待審",
        }.get(cat["hk_tax_deductible"], "-")
        exp_data.append([
            Paragraph(f"{cat['name_zh']}", styles["body"]),
            Paragraph(deduct_label, styles["body"]),
            Paragraph(f"{ccy} {_fmt(cat['total_hkd'])}", styles["right"]),
        ])
    exp_data.append([
        Paragraph("<b>合計支出 / Total Expenses</b>", styles["total_lbl"]),
        Paragraph("", styles["body"]),
        Paragraph(f"<b>{ccy} {_fmt(pnl['total_expense'])}</b>", styles["total_val"]),
    ])
    story.append(make_table(exp_data, [pw * 0.50, pw * 0.15, pw * 0.35]))
    story.append(Spacer(1, 5*mm))

    # ── 利润汇总 ─────────────────────────────────────────────────
    story.append(Paragraph("<b>三、利潤及稅務估算 / Profit & Tax Estimate</b>", styles["section"]))
    story.append(Spacer(1, 2*mm))

    net_color = COLOR_GREEN if pnl["net_profit"] >= 0 else COLOR_RED
    net_s = S("net_val_dyn", fontSize=13, textColor=net_color, leading=18, alignment=TA_RIGHT)

    profit_data = [
        [Paragraph("<b>項目</b>", styles["th"]), Paragraph("<b>金額</b>", styles["th_r"])],
        [Paragraph("毛利潤 / Gross Profit", styles["body"]),
         Paragraph(f"{ccy} {_fmt(pnl['gross_profit'])}", styles["right"])],
        [Paragraph("可稅扣支出 / Tax-Deductible Expenses", styles["body"]),
         Paragraph(f"({ccy} {_fmt(pnl['total_tax_deductible'])})", styles["right"])],
        [Paragraph("應評稅利潤 / Assessable Profit", styles["body"]),
         Paragraph(f"{ccy} {_fmt(pnl['assessable_profit'])}", styles["right"])],
        [Paragraph(f"估算利得稅 / Est. Profits Tax  ({pnl['tax_info']['rate_desc']})", styles["body"]),
         Paragraph(f"({ccy} {_fmt(pnl['tax_info']['estimated_tax'])})", styles["right"])],
        [Paragraph("<b>估算稅後淨利 / Est. Net Profit After Tax</b>", styles["net_lbl"]),
         Paragraph(f"<b>{ccy} {_fmt(pnl['net_profit'])}</b>", net_s)],
    ]
    story.append(make_table(profit_data, [pw * 0.65, pw * 0.35]))
    story.append(Spacer(1, 6*mm))

    # ── 免责声明 ─────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=COLOR_BORDER))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        "免責聲明：本報表由系統根據已錄入數據自動生成，僅供參考，不構成正式財務審計報告。"
        "稅務估算採用香港利得稅兩級稅率制度（法團：首HKD 2,000,000利潤 8.25%，超額部份 16.5%），"
        "實際稅務義務請以香港稅務局評估為準，並建議諮詢持牌會計師。",
        styles["muted"]
    ))

    doc.build(story)
    buffer.seek(0)
    return buffer.read()


# ──────────────────────────────────────────────────────────────
# 5. Excel 导出（损益表）
# ──────────────────────────────────────────────────────────────

def build_pnl_excel(pnl: dict, company_data: dict) -> bytes:
    """生成损益表 Excel，含深蓝标题行、颜色区分可税扣项目。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl 未安装，请运行: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P&L 损益表"

    NAVY   = "1E3A8A"
    LIGHT  = "EFF6FF"
    GREEN  = "DCFCE7"
    RED    = "FEE2E2"
    YELLOW = "FEF9C3"
    WHITE  = "FFFFFF"
    MUTED  = "64748B"

    def cell(row, col, value="", bold=False, bg=None, color="000000",
             align="left", fmt=None, size=10):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, color=color, size=size, name="微软雅黑")
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if fmt:
            c.number_format = fmt
        return c

    ccy = pnl.get("base_currency", "HKD")
    comp = company_data.get("name_zh", "") or company_data.get("name_en", "")
    r = 1

    # 标题
    ws.merge_cells(f"A{r}:D{r}")
    cell(r, 1, f"{comp}  |  損益表 / P&L Statement  |  {pnl['fiscal_year']} 財政年度",
         bold=True, size=14, bg=NAVY, color="FFFFFF", align="center")
    ws.row_dimensions[r].height = 28
    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    cell(r, 1, f"期間 {pnl['period_start']} 至 {pnl['period_end']}    製表：{datetime.utcnow().strftime('%Y-%m-%d')}",
         size=9, color=MUTED, align="center")
    r += 2

    # ── 收入 ──────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:D{r}")
    cell(r, 1, "一、收入 / Income", bold=True, size=11, bg=NAVY, color="FFFFFF")
    ws.row_dimensions[r].height = 20
    r += 1
    # 表头
    for ci, h in enumerate(["項目", "說明", "金額 (HKD)", "佔比"], 1):
        cell(r, ci, h, bold=True, bg="334155", color="FFFFFF", align="center")
    r += 1

    total_inv = pnl["invoice_income"]
    total_com = pnl["commission_income"]
    total_inc = pnl["total_income"]

    cell(r, 1, "發票收入 / Invoice Revenue", bg=WHITE)
    cell(r, 2, "已付款發票", bg=WHITE, color=MUTED)
    cell(r, 3, total_inv, bg=WHITE, align="right", fmt='#,##0.00')
    cell(r, 4, f"{total_inv/total_inc*100:.1f}%" if total_inc else "—", bg=WHITE, align="right")
    r += 1

    src = "IR56M年度報表" if pnl["commission_source"] == "ir56m" else "月結單彙總"
    cell(r, 1, "佣金收入 / Commission Income", bg=LIGHT)
    cell(r, 2, src, bg=LIGHT, color=MUTED)
    cell(r, 3, total_com, bg=LIGHT, align="right", fmt='#,##0.00')
    cell(r, 4, f"{total_com/total_inc*100:.1f}%" if total_inc else "—", bg=LIGHT, align="right")
    r += 1

    cell(r, 1, "合計收入 / Total Income", bold=True, bg=NAVY, color="FFFFFF")
    cell(r, 2, "", bg=NAVY)
    cell(r, 3, total_inc, bold=True, bg=NAVY, color="FFFFFF", align="right", fmt='#,##0.00')
    cell(r, 4, "100%", bold=True, bg=NAVY, color="FFFFFF", align="right")
    r += 2

    # ── 支出 ──────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:D{r}")
    cell(r, 1, "二、支出 / Expenses", bold=True, size=11, bg=NAVY, color="FFFFFF")
    ws.row_dimensions[r].height = 20
    r += 1
    for ci, h in enumerate(["支出類別", "利得稅可扣", "金額 (HKD)", "佔比"], 1):
        cell(r, ci, h, bold=True, bg="334155", color="FFFFFF", align="center")
    r += 1

    DEDUCT_LABEL = {
        "yes": "✓ 全額", "partial": "◑ 部份",
        "depreciation": "↓ 折舊", "no": "✗ 不可", "review": "？待審",
    }
    DEDUCT_BG = {
        "yes": GREEN, "partial": YELLOW,
        "depreciation": YELLOW, "no": RED, "review": WHITE,
    }
    alt = [WHITE, LIGHT]
    total_exp = pnl["total_expense"]

    for i, cat in enumerate(pnl["expense_categories"]):
        bg = DEDUCT_BG.get(cat["hk_tax_deductible"], alt[i % 2])
        cell(r, 1, cat["name_zh"], bg=bg)
        cell(r, 2, DEDUCT_LABEL.get(cat["hk_tax_deductible"], "?"), bg=bg, align="center")
        cell(r, 3, cat["total_hkd"], bg=bg, align="right", fmt='#,##0.00')
        cell(r, 4, f"{cat['total_hkd']/total_exp*100:.1f}%" if total_exp else "—",
             bg=bg, align="right")
        r += 1

    cell(r, 1, "合計支出 / Total Expenses", bold=True, bg=NAVY, color="FFFFFF")
    cell(r, 2, "", bg=NAVY)
    cell(r, 3, total_exp, bold=True, bg=NAVY, color="FFFFFF", align="right", fmt='#,##0.00')
    cell(r, 4, "100%", bold=True, bg=NAVY, color="FFFFFF", align="right")
    r += 2

    # ── 利润 ──────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:D{r}")
    cell(r, 1, "三、利潤及稅務估算 / Profit & Tax", bold=True, size=11, bg=NAVY, color="FFFFFF")
    ws.row_dimensions[r].height = 20
    r += 1

    profit_rows = [
        ("毛利潤 / Gross Profit",                     pnl["gross_profit"],         WHITE),
        ("可稅扣支出 / Tax-Deductible Exp.",           -pnl["total_tax_deductible"], LIGHT),
        ("應評稅利潤 / Assessable Profit",             pnl["assessable_profit"],    YELLOW),
        (f"估算利得稅 ({pnl['tax_info']['rate_desc']})",      -pnl["tax_info"]["estimated_tax"], RED),
        ("估算稅後淨利 / Est. Net Profit After Tax",  pnl["net_profit"],            GREEN),
    ]
    for label, val, bg in profit_rows:
        is_net = "淨利" in label
        cell(r, 1, label, bold=is_net, bg=bg, size=11 if is_net else 10)
        cell(r, 2, ccy, bg=bg, align="right")
        cell(r, 3, val, bold=is_net, bg=bg, align="right", fmt='#,##0.00', size=11 if is_net else 10)
        cell(r, 4, "", bg=bg)
        if is_net:
            ws.row_dimensions[r].height = 22
        r += 1

    r += 1
    cell(r, 1,
         "免責聲明：本表格由系統自動生成，僅供參考。稅務估算採用香港法團利得稅兩級制，"
         "實際稅務義務以稅務局評估為準。",
         size=8, color=MUTED)
    ws.merge_cells(f"A{r}:D{r}")

    # 列宽
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
