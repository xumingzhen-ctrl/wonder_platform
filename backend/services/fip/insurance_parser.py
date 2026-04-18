"""
insurance_parser.py
────────────────────────────────────────────────────────────────────────────
保险计划书 Excel 解析器

支持：香港分红型保险（参与型）计划书标准格式
- 自动识别多级合并表头（兼容 2-4 行表头）
- 关键词匹配中英文列名
- 输出：标准化的年度现金价值数据列表

适配已知格式
  - saving plan cash flow.xlsx  （100年全单，无提取）
  - withdraw plan.xlsx          （多 Tab，含不同提取方案）
────────────────────────────────────────────────────────────────────────────
"""
import io
from typing import Optional
import openpyxl


# ── 列名关键词映射表（中英文均兼容） ───────────────────────────────────────────
_KEYWORD_MAP = {
    'year': [
        '保单年度终结', '保单年度', '年度', 'policy year', 'year', '期末', '年底',
    ],
    'guaranteed_cv': [
        '保证现金价值(a)', '保证现金价值 (a)', '保证现金价值', '保证金额(2)', '保证金额', '保证', 'guaranteed cash value', 'guaranteed cv',
        '保证值', '(a)', 'gcv', '基本现金价值',
    ],
    'rev_bonus': [
        '复归红利 (b)', '复归红利', '可归红利', 'reversionary bonus', 'rev bonus', 'rev. bonus',
        '(b)', '复归', '已宣告红利',
    ],
    'terminal_bonus': [
        '终期分红 (c)', '终期分红', '终期红利', 'terminal bonus', '(c)', '终期',
        'terminal dividend', '特别红利',
    ],
    'withdrawal': [
        '提取金额', '现金提取金额(1),(2)', '现金提取金额', '现金提取', 'withdrawal', 'cash withdrawal',
        '每年提取', '年提取', '退保', '部分退保',
    ],
    'premium': [
        '保费', '目标保费', '已缴保费', '保费金额', '总缴保费', '年缴保费', '缴付保费', 'premium', 'annual premium', 'paid premium'
    ],
    'total_cv': [
        '总额 (a)+(b)+(c)', '总额', '(a)+(b)+(c)', '总现金价值', 'total', 'total cv',
        '合计', '累计现金价值', '总现值',
    ],
}


def _match_column(text: str) -> Optional[str]:
    """
    将合并后的表头文字与关键词映射对比，返回字段名；无匹配返回 None。
    策略：忽略大小写、忽略空格，取最长匹配关键词。
    """
    if not text:
        return None
    cleaned = str(text).lower().replace(' ', '').replace('\n', '')
    for field, keywords in _KEYWORD_MAP.items():
        for kw in keywords:
            if kw.lower().replace(' ', '') in cleaned:
                # 排除将“已缴保费总额”错误识别为“总现金价值”的情况
                if field == 'total_cv' and ('保费' in cleaned or 'premium' in cleaned):
                    continue
                return field
    return None


def _read_headers(ws, max_header_rows: int = 4) -> dict:
    """
    扫描前 max_header_rows 行，合并每列的文字内容，
    返回 {列索引(1-based): 字段名} 的映射。
    """
    col_texts = {}  # type: dict[int, str]

    # 收集各列的表头文字（多行合并）
    for row_idx in range(1, max_header_rows + 1):
        for cell in ws[row_idx]:
            col = cell.column
            val = cell.value
            if val is not None:
                col_texts[col] = col_texts.get(col, '') + str(val)

    # 关键词匹配
    col_map = {}  # type: dict[int, str]
    for col, text in col_texts.items():
        field = _match_column(text)
        if field and field not in col_map.values():
            col_map[col] = field

    return col_map


def _find_data_start_row(ws, year_col: int, max_scan: int = 15) -> int:
    """
    从第2行开始扫描，找到 year_col 列出现整数 1 的那一行，即数据起始行。
    """
    for row_idx in range(2, max_scan + 2):
        cell_val = ws.cell(row=row_idx, column=year_col).value
        if cell_val is not None:
            try:
                num = float(str(cell_val).replace(',', ''))
                if int(num) == 1:
                    return row_idx
            except (ValueError, TypeError):
                continue
    return 5  # 默认第5行（withdraw plan.xlsx 格式）


def parse_insurance_plan(file_bytes: bytes, filename: str) -> dict:
    """
    解析保险计划书 Excel 文件，返回标准化的年度数据。

    Parameters
    ----------
    file_bytes : bytes
        Excel 文件的二进制内容
    filename : str
        原始文件名（用于推断 policy_name）

    Returns
    -------
    dict
        成功时：
        {
            "ok": True,
            "policy_name": str,
            "total_years": int,
            "years": [
                {
                    "year": int,
                    "guaranteed_cv": float,
                    "rev_bonus": float,
                    "terminal_bonus": float,
                    "non_guaranteed": float,   # = rev_bonus + terminal_bonus
                    "withdrawal": float,        # 0.0 如无提取列
                    "total_cv_base": float      # 保险公司给出的基准总CV
                },
                ...
            ]
        }
        失败时：
        {
            "ok": False,
            "error": str
        }
    """
    # ── 1. 打开 Excel ──────────────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return {"ok": False, "error": f"无法打开 Excel 文件：{e}"}

    # 取第一个工作表（用户每次只上传一个 Tab 的数据）
    ws = wb.active

    # ── 2. 识别表头 ────────────────────────────────────────────────────────────
    col_map = _read_headers(ws, max_header_rows=4)

    if 'year' not in col_map.values():
        # 尝试更宽泛的扫描（前6行）
        col_map = _read_headers(ws, max_header_rows=6)

    if 'year' not in col_map.values():
        return {
            "ok": False,
            "error": (
                "未能识别「保单年度」列。"
                "请确认上传的是保险计划书 Excel，"
                "且包含「保单年度」、「保证现金价值」等标准列名。"
            )
        }

    # 建立字段 → 列索引的反向映射
    field_to_col = {v: k for k, v in col_map.items()}  # type: dict

    year_col     = field_to_col['year']
    gcv_col      = field_to_col.get('guaranteed_cv')
    rev_col      = field_to_col.get('rev_bonus')
    term_col     = field_to_col.get('terminal_bonus')
    total_col    = field_to_col.get('total_cv')
    wdraw_col    = field_to_col.get('withdrawal')
    prem_col     = field_to_col.get('premium')

    # ── 3. 识别数据起始行 ──────────────────────────────────────────────────────
    data_start = _find_data_start_row(ws, year_col, max_scan=12)

    # ── 4. 逐行读取数据 ────────────────────────────────────────────────────────
    years_data = []
    prev_year = 0

    for row_idx in range(data_start, ws.max_row + 1):
        year_val = ws.cell(row=row_idx, column=year_col).value

        # 年度列为空 → 停止读取
        if year_val is None:
            break

        try:
            year = int(float(str(year_val)))
        except (ValueError, TypeError):
            continue

        # 单调性检查（跳过异常行）
        if year <= prev_year:
            continue
        prev_year = year

        def get_float(col):
            if col is None:
                return 0.0
            val = ws.cell(row=row_idx, column=col).value
            if val is None:
                return 0.0
            try:
                return float(str(val).replace(',', ''))
            except (ValueError, TypeError):
                return 0.0

        gcv        = abs(get_float(gcv_col))
        rev        = abs(get_float(rev_col))
        terminal   = abs(get_float(term_col))
        total      = abs(get_float(total_col))
        withdrawal = abs(get_float(wdraw_col))
        premium    = abs(get_float(prem_col))

        non_guaranteed = rev + terminal

        # 如果 total_cv 列未识别，自动计算
        if total == 0.0 and (gcv > 0 or non_guaranteed > 0):
            total = gcv + non_guaranteed

        years_data.append({
            "year":           year,
            "premium":        round(premium, 2),
            "guaranteed_cv":  round(gcv, 2),
            "rev_bonus":      round(rev, 2),
            "terminal_bonus": round(terminal, 2),
            "non_guaranteed": round(non_guaranteed, 2),
            "withdrawal":     round(withdrawal, 2),
            "total_cv_base":  round(total, 2),
        })

    if not years_data:
        return {
            "ok": False,
            "error": "未能从文件中读取任何有效数据行。请检查文件格式或数据列是否正确。"
        }

    # ── 6. 额外处理：推导年度保费 ───────────────────────────────────────
    # 如果 premium 是单调不减且最后一年的 premium 等于第一年但不为0，或者连续增长，可能是累计保费
    if len(years_data) > 1:
        is_cumulative = False
        p0 = years_data[0]['premium']
        p1 = years_data[1]['premium']
        
        # 判断累计：第二年保费 > 第一年（如 100k > 50k），或前两年相等但大于0（极少情况，或者就是一次性趸交这种也无所谓）
        if p1 > p0:
            is_cumulative = True
            
        if is_cumulative:
            last_p = 0
            for r in years_data:
                true_p = r['premium'] - last_p
                # 防止由于数据瑕疵导致的负数
                r['premium'] = max(0.0, true_p)
                last_p += r['premium']

    # ── 5. 验证：至少有保证CV或总CV列 ─────────────────────────────────────────
    has_meaningful_data = any(
        r['guaranteed_cv'] > 0 or r['total_cv_base'] > 0
        for r in years_data
    )
    if not has_meaningful_data:
        return {
            "ok": False,
            "error": "已读取数据，但所有现金价值均为 0。请确认选择的 Sheet 包含现金价值数据。"
        }

    # ── 6. 推断 policy_name ───────────────────────────────────────────────────
    policy_name = filename.rsplit('.', 1)[0] if filename else "保险计划书"

    return {
        "ok":          True,
        "policy_name": policy_name,
        "total_years": len(years_data),
        "years":       years_data,
    }


# ── 简单命令行测试入口 ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    import json

    if len(sys.argv) < 2:
        print("Usage: python insurance_parser.py <path_to_excel>")
        sys.exit(1)

    path = sys.argv[1]
    with open(path, 'rb') as f:
        data = f.read()

    result = parse_insurance_plan(data, path.split('/')[-1])
    if result['ok']:
        print(f"✅ 解析成功：{result['policy_name']}，共 {result['total_years']} 年")
        print("前5年数据：")
        for row in result['years'][:5]:
            print(f"  Year{row['year']:3d}: GCV={row['guaranteed_cv']:10,.0f} "
                  f"非保证={row['non_guaranteed']:10,.0f} "
                  f"提取={row['withdrawal']:8,.0f} "
                  f"总CV={row['total_cv_base']:10,.0f}")
    else:
        print(f"❌ 解析失败：{result['error']}")
