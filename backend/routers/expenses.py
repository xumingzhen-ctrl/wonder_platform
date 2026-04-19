"""
expenses.py — 支出凭证管理 API 路由

提供收据上传、查询、人工复核、统计分析等接口。
"""
import asyncio
import logging
from datetime import date
from typing import List, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, BackgroundTasks
from sqlalchemy import func
from sqlalchemy.orm import Session

from database import get_db
from models.company import User, UserCompanyAccess
from models.expense import Expense, ExpenseCategory, ExpenseStatus, ReceiptType
from services.auth import get_current_user
from services.receipt_scanner import scan_receipt
from config import settings, ensure_receipt_dirs
from fastapi.responses import StreamingResponse
import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/expenses", tags=["支出管理"])

def _check_company_access(company_id: str, user: User, db: Session):
    access = db.query(UserCompanyAccess).filter(
        UserCompanyAccess.company_id == company_id,
        UserCompanyAccess.user_id == user.id
    ).first()
    if not access:
        raise HTTPException(status_code=403, detail="无此公司访问权限")

# ── 种子数据注入（在分类表为空时自动初始化）────────────────────────

EXPENSE_CATEGORIES_SEED = [
    {"code": "MEAL",          "name_zh": "餐饮招待",   "name_en": "Meals & Entertainment (Business)", "hk_tax_deductible": "partial",     "hk_tax_note": "商务餐饮50%可抵扣，私人性质不可扣", "sort_order": 1},
    {"code": "TRAVEL",        "name_zh": "差旅交通",   "name_en": "Travel & Transportation",          "hk_tax_deductible": "yes",          "hk_tax_note": "商务差旅全额可扣，须保留单据", "sort_order": 2},
    {"code": "OFFICE",        "name_zh": "办公耗材",   "name_en": "Office Supplies",                  "hk_tax_deductible": "yes",          "hk_tax_note": "文具、打印耗材等全额可扣", "sort_order": 3},
    {"code": "SOFTWARE",      "name_zh": "软件订阅",    "name_en": "Software & SaaS Subscriptions",   "hk_tax_deductible": "yes",          "hk_tax_note": "商业用途SaaS/云服务/软件使用权费用全额可扣，须保留发票", "sort_order": 4},
    {"code": "MARKETING",     "name_zh": "市场营销",   "name_en": "Marketing & Advertising",          "hk_tax_deductible": "yes",          "hk_tax_note": "广告、印刷、推广费用全额可扣", "sort_order": 6},
    {"code": "STAFF",         "name_zh": "员工津贴福利","name_en": "Staff Benefits & Allowances",      "hk_tax_deductible": "yes",          "hk_tax_note": "员工薪酬、福利全额可扣", "sort_order": 5},
    {"code": "PROFESSIONAL",  "name_zh": "专业服务费", "name_en": "Professional Fees",                "hk_tax_deductible": "yes",          "hk_tax_note": "律师费、会计师费、顾问费全额可扣", "sort_order": 6},
    {"code": "EQUIPMENT",     "name_zh": "设备采购",   "name_en": "Equipment & Machinery",            "hk_tax_deductible": "depreciation", "hk_tax_note": "资本性支出按折旧摊销，电脑等可申请一次性抵扣", "sort_order": 7},
    {"code": "UTILITIES",     "name_zh": "水电通讯",   "name_en": "Utilities & Communications",       "hk_tax_deductible": "yes",          "hk_tax_note": "办公场所水电、宽带、电话按商用比例可扣", "sort_order": 8},
    {"code": "RENT",          "name_zh": "租金",       "name_en": "Rent",                             "hk_tax_deductible": "yes",          "hk_tax_note": "商业用途租金全额可扣", "sort_order": 9},
    {"code": "INSURANCE",     "name_zh": "保险",       "name_en": "Insurance",                        "hk_tax_deductible": "yes",          "hk_tax_note": "商业保险（非人寿险）全额可扣", "sort_order": 10},
    {"code": "ENTERTAINMENT", "name_zh": "商务娱乐",   "name_en": "Business Entertainment",           "hk_tax_deductible": "no",           "hk_tax_note": "按香港税务局裁定，商务娱乐一律不可扣", "sort_order": 11},
    {"code": "OTHER",         "name_zh": "其他",       "name_en": "Other",                            "hk_tax_deductible": "review",       "hk_tax_note": "需人工确认是否符合抵扣条件", "sort_order": 12},
]


def _ensure_categories(db: Session):
    """如果分类表为空，自动注入种子数据"""
    count = db.query(func.count(ExpenseCategory.id)).scalar()
    if count == 0:
        for cat_data in EXPENSE_CATEGORIES_SEED:
            cat = ExpenseCategory(**cat_data)
            db.add(cat)
        db.commit()
        logger.info(f"已初始化 {len(EXPENSE_CATEGORIES_SEED)} 个支出分类")


# ── API Endpoints ─────────────────────────────────────────────────

@router.get("/categories")
def list_categories(db: Session = Depends(get_db)):
    """获取所有支出分类（含香港利得税抵扣信息）"""
    _ensure_categories(db)
    cats = db.query(ExpenseCategory).filter(
        ExpenseCategory.is_active == True
    ).order_by(ExpenseCategory.sort_order).all()

    return [
        {
            "id": c.id,
            "code": c.code,
            "name_zh": c.name_zh,
            "name_en": c.name_en,
            "hk_tax_deductible": c.hk_tax_deductible,
            "hk_tax_note": c.hk_tax_note,
        }
        for c in cats
    ]


@router.post("/upload")
async def upload_receipts(
    background_tasks: BackgroundTasks,
    company_id: str = Form(...),
    files: List[UploadFile] = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    批量上传收据图片（支持 JPG/PNG/HEIC/PDF）

    - 每个文件独立处理，单个失败不影响其他文件
    - AI 识别完成后记录状态为 pending，等待人工复核
    """
    if current_user.role not in ["admin", "premium"]:
        from datetime import date
        import calendar
        today = date.today()
        first_day = today.replace(day=1)
        last_day = today.replace(day=calendar.monthrange(today.year, today.month)[1])
        
        count = db.query(Expense).filter(
            Expense.company_id == company_id,
            Expense.created_at >= first_day,
            Expense.created_at <= last_day,
            Expense.ai_confidence.isnot(None)
        ).count()
        
        if count + len(files) > 10:
            raise HTTPException(status_code=403, detail=f"免費版帳戶每月最多自動識別 10 張發票。本月已使用 {count} 張。如需處理更多，請升級帳戶。")

    ensure_receipt_dirs()
    _ensure_categories(db)

    results = []

    for upload_file in files:
        filename = upload_file.filename or "unknown"
        try:
            file_bytes = await upload_file.read()
            extracted = await scan_receipt(file_bytes, filename, db, company_id)

            # 查找对应分类
            category = db.query(ExpenseCategory).filter(
                ExpenseCategory.code == extracted["suggested_category_code"]
            ).first()

            expense = Expense(
                company_id=company_id,
                voucher_number=extracted["voucher_number"],
                receipt_type=extracted.get("receipt_type"),
                receipt_date=extracted.get("receipt_date"),
                vendor_name=extracted.get("vendor_name"),
                vendor_tax_id=extracted.get("vendor_tax_id"),
                description=extracted.get("description"),
                currency=extracted.get("currency", "HKD"),
                amount_original=extracted.get("amount_original"),
                tax_rate=extracted.get("tax_rate"),
                tax_amount=extracted.get("tax_amount"),
                total_amount=extracted.get("total_amount"),
                amount_hkd=extracted.get("amount_hkd"),
                cn_invoice_code=extracted.get("cn_invoice_code"),
                cn_invoice_number=extracted.get("cn_invoice_number"),
                category_id=category.id if category else None,
                ai_confidence=extracted.get("ai_confidence"),
                ai_raw_response=extracted.get("ai_raw_response"),
                receipt_image_path=extracted.get("receipt_image_path"),
                receipt_original_filename=extracted.get("receipt_original_filename"),
                source_format=extracted.get("source_format"),
                fiscal_year=extracted.get("fiscal_year"),
                file_hash=extracted.get("file_hash"),
                status=ExpenseStatus.pending,
            )
            db.add(expense)
            db.commit()
            db.refresh(expense)

            results.append({
                "filename": filename,
                "status": "success",
                "voucher_number": expense.voucher_number,
                "vendor_name": expense.vendor_name,
                "total_amount": float(expense.total_amount) if expense.total_amount else None,
                "amount_hkd": float(expense.amount_hkd) if expense.amount_hkd else None,
                "currency": expense.currency,
                "receipt_date": str(expense.receipt_date) if expense.receipt_date else None,
                "fiscal_year": expense.fiscal_year,
                "ai_confidence": expense.ai_confidence,
                "expense_id": expense.id,
            })

        except Exception as e:
            logger.error(f"处理文件失败 [{filename}]: {e}", exc_info=True)
            results.append({
                "filename": filename,
                "status": "error",
                "error": str(e),
            })

    success_count = sum(1 for r in results if r["status"] == "success")
    return {
        "total": len(files),
        "success": success_count,
        "failed": len(files) - success_count,
        "results": results,
    }

@router.post("/manual")
async def add_expense_manual(
    company_id: str = Form(...),
    receipt_date: str = Form(...),
    vendor_name: str = Form(...),
    total_amount: float = Form(...),
    category_code: str = Form(...),
    currency: str = Form("HKD"),
    description: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """手动记账（无收据图片）"""
    _check_company_access(company_id, current_user, db)
    _ensure_categories(db)
    
    category = db.query(ExpenseCategory).filter(ExpenseCategory.code == category_code).first()
    
    from services.receipt_scanner import calculate_fiscal_year, convert_to_hkd
    from datetime import datetime
    
    try:
        parsed_date = datetime.strptime(receipt_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="无效的日期格式，需为 YYYY-MM-DD")
        
    fiscal_year = calculate_fiscal_year(parsed_date)
    amount_hkd = convert_to_hkd(total_amount, currency)
    
    count = db.query(Expense).filter(
        Expense.company_id == company_id,
        Expense.receipt_date == parsed_date
    ).count() + 1
    import uuid
    short_uuid = str(uuid.uuid4())[:6].upper()
    voucher_number = f"EXP-{parsed_date.strftime('%Y%m')}-M{count:03d}-{short_uuid}"
    
    expense = Expense(
        company_id=company_id,
        voucher_number=voucher_number,
        receipt_date=parsed_date,
        vendor_name=vendor_name,
        description=description,
        currency=currency,
        total_amount=total_amount,
        amount_hkd=amount_hkd,
        category_id=category.id if category else None,
        fiscal_year=fiscal_year,
        status=ExpenseStatus.confirmed, # 手动录入直接标记为已确认
        notes=notes,
        receipt_type="other"
    )
    
    db.add(expense)
    db.commit()
    db.refresh(expense)
    
    return _expense_to_dict(expense, db)



@router.post("/scan-inbox")
async def scan_inbox(
    company_id: Optional[str] = Form(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    触发 inbox 文件夹全量扫描（用于历史数据批量导入）

    如未指定 company_id，使用 .env 中的 DEFAULT_COMPANY_ID
    """
    from services.inbox_watcher import run_scan_once

    target_company = company_id or settings.DEFAULT_COMPANY_ID
    if not target_company:
        raise HTTPException(
            status_code=400,
            detail="请提供 company_id 或在 .env 中配置 DEFAULT_COMPANY_ID"
        )
        
    if current_user.role not in ["admin", "premium"]:
        from datetime import date
        import calendar
        import os
        from config import settings
        
        today = date.today()
        first_day = today.replace(day=1)
        last_day = today.replace(day=calendar.monthrange(today.year, today.month)[1])
        
        count = db.query(Expense).filter(
            Expense.company_id == target_company,
            Expense.created_at >= first_day,
            Expense.created_at <= last_day,
            Expense.ai_confidence.isnot(None)
        ).count()
        
        # Check files in inbox
        inbox_dir = settings.RECEIPTS_INBOX_PATH
        files_count = len([f for f in os.listdir(inbox_dir) if not f.startswith(".")]) if os.path.exists(inbox_dir) else 0
        
        if count + files_count > 10:
            raise HTTPException(status_code=403, detail=f"免費版帳戶每月最多自動識別 10 張發票。本月已识别 {count} 张，收件箱还有 {files_count} 张等待处理，超额。请升级。")

    result = await run_scan_once(target_company, db)
    return result


@router.get("/")
def list_expenses(
    company_id: str = Query(...),
    fiscal_year: Optional[str] = Query(None, description="如：2023-24"),
    category_code: Optional[str] = Query(None),
    status: Optional[str] = Query(None, description="pending|confirmed|rejected"),
    receipt_type: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """支出记录列表查询（支持多维筛选）"""
    _check_company_access(company_id, current_user, db)
    query = db.query(Expense).filter(Expense.company_id == company_id)

    if fiscal_year:
        query = query.filter(Expense.fiscal_year == fiscal_year)
    if status:
        query = query.filter(Expense.status == status)
    if receipt_type:
        query = query.filter(Expense.receipt_type == receipt_type)
    if date_from:
        query = query.filter(Expense.receipt_date >= date_from)
    if date_to:
        query = query.filter(Expense.receipt_date <= date_to)
    if category_code:
        cat = db.query(ExpenseCategory).filter(ExpenseCategory.code == category_code).first()
        if cat:
            query = query.filter(Expense.category_id == cat.id)

    total = query.count()
    expenses = (
        query.order_by(Expense.receipt_date.desc().nulls_last(), Expense.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [_expense_to_dict(e, db) for e in expenses],
    }


@router.get("/stats/by-category")
def stats_by_category(
    company_id: str = Query(...),
    fiscal_year: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """按分类的支出统计"""
    _check_company_access(company_id, current_user, db)
    query = db.query(
        ExpenseCategory.code,
        ExpenseCategory.name_zh,
        ExpenseCategory.hk_tax_deductible,
        func.count(Expense.id).label("count"),
        func.sum(Expense.amount_hkd).label("total_hkd"),
    ).join(Expense, Expense.category_id == ExpenseCategory.id, isouter=True).filter(
        Expense.company_id == company_id,
        Expense.status != ExpenseStatus.rejected,
    )

    if fiscal_year:
        query = query.filter(Expense.fiscal_year == fiscal_year)

    results = query.group_by(ExpenseCategory.code, ExpenseCategory.name_zh, ExpenseCategory.hk_tax_deductible).all()

    return [
        {
            "category_code": r.code,
            "category_name": r.name_zh,
            "hk_tax_deductible": r.hk_tax_deductible,
            "count": r.count or 0,
            "total_hkd": float(r.total_hkd or 0),
        }
        for r in results
    ]


@router.get("/stats/by-fiscal-year")
def stats_by_fiscal_year(
    company_id: str = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """按财政年度的支出汇总"""
    _check_company_access(company_id, current_user, db)
    results = db.query(
        Expense.fiscal_year,
        func.count(Expense.id).label("count"),
        func.sum(Expense.amount_hkd).label("total_hkd"),
    ).filter(
        Expense.company_id == company_id,
        Expense.status != ExpenseStatus.rejected,
        Expense.fiscal_year.isnot(None),
    ).group_by(Expense.fiscal_year).order_by(Expense.fiscal_year.desc()).all()

    return [
        {
            "fiscal_year": r.fiscal_year,
            "count": r.count,
            "total_hkd": float(r.total_hkd or 0),
        }
        for r in results
    ]



@router.get("/export")
def export_expenses(
    company_id: str = Query(...),
    format: str = Query("excel", regex="^(excel|csv)$"),
    status: Optional[str] = None,
    fiscal_year: Optional[str] = None,
    category_code: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    导出支出明细为 Excel (.xlsx) 或 CSV 文件。
    支持与列表页相同的筛选条件。
    """
    _check_company_access(company_id, current_user, db)
    # ── 查询 ──────────────────────────────────────────────────────
    q = db.query(Expense).filter(Expense.company_id == company_id)
    if status:
        try:
            q = q.filter(Expense.status == ExpenseStatus[status])
        except KeyError:
            pass
    if fiscal_year:
        q = q.filter(Expense.fiscal_year == fiscal_year)
    if category_code:
        q = q.join(ExpenseCategory, Expense.category_id == ExpenseCategory.id, isouter=True)
        q = q.filter(ExpenseCategory.code == category_code)
    expenses = q.order_by(Expense.receipt_date.desc()).all()

    # ── 列定义 ───────────────────────────────────────────────────
    HEADERS = [
        ("凭证号",         lambda e: e.voucher_number or ""),
        ("收据日期",       lambda e: str(e.receipt_date) if e.receipt_date else ""),
        ("财政年度",       lambda e: e.fiscal_year or ""),
        ("商户名称",       lambda e: e.vendor_name or ""),
        ("发票类型",       lambda e: {
            "cn_vat_special": "增值税专用发票",
            "cn_vat_general": "增值税普通发票",
            "cn_ordinary": "内地普通收据",
            "hk_receipt": "香港收据",
        }.get(str(e.receipt_type).replace("ReceiptType.", ""), "其他")),
        ("货币",           lambda e: e.currency or ""),
        ("原始金额",       lambda e: float(e.total_amount) if e.total_amount else ""),
        ("港元金额(HKD)", lambda e: float(e.amount_hkd) if e.amount_hkd else ""),
        ("支出分类",       lambda e: e.category.name_zh if e.category else "未分类"),
        ("港税可扣",       lambda e: {
            "yes": "全额可扣", "partial": "部分可扣",
            "no": "不可扣", "depreciation": "折旧摊销", "review": "待确认",
        }.get(e.category.hk_tax_deductible if e.category else "", "—")),
        ("AI置信度(%)",   lambda e: e.ai_confidence or ""),
        ("状态",           lambda e: {
            "pending": "待复核", "confirmed": "已确认", "rejected": "已驳回",
        }.get(str(e.status).replace("ExpenseStatus.", ""), "")),
        ("摘要",           lambda e: e.description or ""),
        ("备注",           lambda e: e.notes or ""),
        ("供应商税号",     lambda e: e.vendor_tax_id or ""),
        ("发票号码",       lambda e: e.cn_invoice_number or ""),
        ("原始文件名",     lambda e: e.receipt_original_filename or ""),
    ]

    rows = [[fn(e) for _, fn in HEADERS] for e in expenses]

    # ── CSV 导出 ─────────────────────────────────────────────────
    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([h for h, _ in HEADERS])
        writer.writerows(rows)
        output.seek(0)
        filename = f"expenses_export.csv"
        return StreamingResponse(
            iter([output.getvalue().encode("utf-8-sig")]),  # utf-8-sig = Excel 可识别 BOM
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8\'\'{filename}"}
        )

    # ── Excel 导出 ───────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "支出明细"

    # 标题行样式
    HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
    HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10, name="Arial")
    CENTER       = Alignment(horizontal="center", vertical="center")
    WRAP         = Alignment(wrap_text=True, vertical="center")
    THIN         = Side(style="thin", color="CCCCCC")
    BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # 写标题
    for col_idx, (header, _) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[1].height = 22

    # 写数据行
    STATUS_FILL = {
        "已确认": PatternFill("solid", fgColor="D1FAE5"),
        "已驳回": PatternFill("solid", fgColor="FEE2E2"),
        "待复核": PatternFill("solid", fgColor="FEF9C3"),
    }
    TAX_FILL = {
        "全额可扣": PatternFill("solid", fgColor="D1FAE5"),
        "部分可扣": PatternFill("solid", fgColor="FEF9C3"),
        "不可扣":   PatternFill("solid", fgColor="FEE2E2"),
    }

    for row_idx, row_data in enumerate(rows, start=2):
        status_val = row_data[11]   # 状态列
        tax_val    = row_data[9]    # 港税可扣列
        row_fill   = STATUS_FILL.get(status_val)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            cell.alignment = WRAP
            cell.font = Font(size=10, name="Arial")
            # 金额列右对齐
            if col_idx in (7, 8):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(value, float):
                    cell.number_format = '#,##0.00'
            # 状态列加色
            if col_idx == 12 and row_fill:
                cell.fill = row_fill
            # 港税列加色
            if col_idx == 10:
                tax_f = TAX_FILL.get(str(value))
                if tax_f:
                    cell.fill = tax_f

        ws.row_dimensions[row_idx].height = 18

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动列宽
    COL_WIDTHS = [14, 12, 10, 22, 14, 6, 12, 14, 14, 10, 10, 8, 30, 20, 16, 16, 22]
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 汇总行（最底部）
    if rows:
        sum_row = len(rows) + 2
        ws.cell(row=sum_row, column=1, value="合计").font = Font(bold=True, size=10)
        # 原始金额列（7）和 HKD 金额列（8）求和（仅数字行）
        hkd_sum = sum(r[7] for r in rows if isinstance(r[7], float))
        cell_sum = ws.cell(row=sum_row, column=8, value=hkd_sum)
        cell_sum.number_format = '#,##0.00'
        cell_sum.font = Font(bold=True, size=10)
        cell_sum.fill = PatternFill("solid", fgColor="DBEAFE")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fiscal_tag = f"_{fiscal_year}" if fiscal_year else ""
    filename = f"expenses{fiscal_tag}_export.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8\'\'{filename}"}
    )


@router.get("/{expense_id}")
def get_expense(
    expense_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取单条支出记录详情"""
    expense = db.query(Expense).filter(Expense.id == expense_id).first()
    if not expense:
        raise HTTPException(status_code=404, detail="记录不存在")
    _check_company_access(expense.company_id, current_user, db)
    return _expense_to_dict(expense, db)


@router.put("/{expense_id}")
def update_expense(
    expense_id: str,
    vendor_name: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    receipt_date: Optional[date] = Form(None),
    currency: Optional[str] = Form(None),
    total_amount: Optional[float] = Form(None),
    category_code: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """人工修正 AI 识别结果"""
    expense = db.query(Expense).filter(Expense.id == expense_id).first()
    if not expense:
        raise HTTPException(status_code=404, detail="记录不存在")
    _check_company_access(expense.company_id, current_user, db)

    if vendor_name is not None:
        expense.vendor_name = vendor_name
    if description is not None:
        expense.description = description
    if receipt_date is not None:
        expense.receipt_date = receipt_date
        expense.fiscal_year = None  # 重新计算财年（由触发器处理或手动更新）
    if currency is not None:
        expense.currency = currency
    if total_amount is not None:
        expense.total_amount = total_amount
    
    # 只要货币或金额发生变化，就重算 HKD 金额
    if currency is not None or total_amount is not None:
        from services.receipt_scanner import convert_to_hkd
        # 使用更新后的值（不论是来自 Form 还是数据库原有的）
        expense.amount_hkd = convert_to_hkd(expense.total_amount, expense.currency)

    if notes is not None:
        expense.notes = notes
    if category_code:
        cat = db.query(ExpenseCategory).filter(ExpenseCategory.code == category_code).first()
        if cat:
            expense.category_id = cat.id

    db.commit()
    db.refresh(expense)
    return _expense_to_dict(expense, db)


@router.put("/{expense_id}/confirm")
def confirm_expense(
    expense_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """确认支出记录（pending → confirmed）"""
    expense = db.query(Expense).filter(Expense.id == expense_id).first()
    if not expense:
        raise HTTPException(status_code=404, detail="记录不存在")
    _check_company_access(expense.company_id, current_user, db)

    expense.status = ExpenseStatus.confirmed
    db.commit()
    return {"message": "已确认", "voucher_number": expense.voucher_number}


@router.put("/{expense_id}/reject")
def reject_expense(
    expense_id: str,
    reason: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """驳回支出记录（识别错误/重复）"""
    expense = db.query(Expense).filter(Expense.id == expense_id).first()
    if not expense:
        raise HTTPException(status_code=404, detail="记录不存在")
    _check_company_access(expense.company_id, current_user, db)

    expense.status = ExpenseStatus.rejected
    if reason:
        expense.notes = f"[驳回原因] {reason}\n" + (expense.notes or "")
    db.commit()
    return {"message": "已驳回", "voucher_number": expense.voucher_number}


@router.delete("/{expense_id}")
def delete_expense(
    expense_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """彻底删除支出记录及其文件记录"""
    expense = db.query(Expense).filter(Expense.id == expense_id).first()
    if not expense:
        raise HTTPException(status_code=404, detail="记录不存在")
    _check_company_access(expense.company_id, current_user, db)

    # 物理删除凭证文件
    if expense.receipt_image_path:
        from config import settings
        from pathlib import Path
        # The path was stored relative like "./receipts_archive/..." or "receipts_archive/..."
        # If it's an absolute path or relative, resolving it safely
        fp = Path(expense.receipt_image_path)
        if fp.exists():
            fp.unlink(missing_ok=True)
    
    db.delete(expense)
    db.commit()
    return {"message": "已删除"}


# ── 辅助函数 ─────────────────────────────────────────────────────

def _expense_to_dict(expense: Expense, db: Session) -> dict:
    category = db.query(ExpenseCategory).filter(
        ExpenseCategory.id == expense.category_id
    ).first() if expense.category_id else None

    return {
        "id": expense.id,
        "voucher_number": expense.voucher_number,
        "receipt_type": expense.receipt_type,
        "receipt_date": str(expense.receipt_date) if expense.receipt_date else None,
        "fiscal_year": expense.fiscal_year,
        "vendor_name": expense.vendor_name,
        "vendor_tax_id": expense.vendor_tax_id,
        "description": expense.description,
        "currency": expense.currency,
        "amount_original": float(expense.amount_original) if expense.amount_original else None,
        "tax_rate": float(expense.tax_rate) if expense.tax_rate else None,
        "tax_amount": float(expense.tax_amount) if expense.tax_amount else None,
        "total_amount": float(expense.total_amount) if expense.total_amount else None,
        "amount_hkd": float(expense.amount_hkd) if expense.amount_hkd else None,
        "cn_invoice_code": expense.cn_invoice_code,
        "cn_invoice_number": expense.cn_invoice_number,
        "category": {
            "code": category.code,
            "name_zh": category.name_zh,
            "hk_tax_deductible": category.hk_tax_deductible,
            "hk_tax_note": category.hk_tax_note,
        } if category else None,
        "ai_confidence": expense.ai_confidence,
        "receipt_image_path": expense.receipt_image_path,
        "receipt_original_filename": expense.receipt_original_filename,
        "source_format": expense.source_format,
        "status": expense.status,
        "notes": expense.notes,
        "upload_date": expense.upload_date.isoformat() if expense.upload_date else None,
        "created_at": expense.created_at.isoformat() if expense.created_at else None,
    }
